# GAP 8.G — EDD EXCEL IMPORT — services/edd_excel_import.py
"""AI-powered EDD Excel import service.

Handles Excel files with arbitrary column formats by using LLM to map
columns to the canonical EDD schema (REUnit fields).

Flow:
  1. Read Excel file with openpyxl
  2. Extract headers + sample rows
  3. Send to LLM for column mapping (AI agent)
  4. Apply mapping to create/update REProject, REBlock, RELevel, REUnit
  5. Return import summary with any warnings
"""
from __future__ import annotations

import io
import json
import uuid
from typing import Any

import structlog
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.bim_edd import (
    REProject,
    REBlock,
    RELevel,
    REUnit,
    EddState,
)

logger = structlog.get_logger(__name__)

# ── Canonical fields that we expect the LLM to map to ───────────────────
CANONICAL_FIELDS = {
    "project_code": "Code projet (ex: PROJ01)",
    "block_code": "Code bloc/immeuble (ex: A, B, BLOC1)",
    "level_code": "Code niveau/etage (ex: L01, RDC, SS1, 3eme)",
    "unit_code": "Code lot/unite unique (ex: A-L01-001)",
    "typology": "Typologie (ex: F2, F3, STUDIO, DUPLEX, LOCAL_COMMERCIAL)",
    "bedrooms": "Nombre de chambres",
    "bathrooms": "Nombre de salles de bain",
    "area_sh": "Surface habitable (m2)",
    "area_su": "Surface utile (m2)",
    "area_sb": "Surface balcon (m2)",
    "area_st": "Surface terrasse (m2)",
    "area_sj": "Surface jardin (m2)",
    "area_sp": "Surface parking (m2)",
    "area_sc": "Surface cave (m2)",
    "exposure": "Orientation/exposition (N, NE, E, SE, S, SO, O, NO)",
    "view_class": "Type de vue (mer, ville, jardin, montagne, none)",
    "price_base": "Prix de base (DZD ou DZD/m2)",
    "price_total": "Prix total (DZD)",
    "usage_type": "Type d'usage (habitation, commerce, bureau, parking)",
}

COLUMN_MAPPING_PROMPT = """Tu es un expert en immobilier algerien. On t'envoie les en-tetes d'un fichier Excel EDD (Etat Descriptif de Division).

Tu dois mapper chaque en-tete vers le champ canonique correspondant. Les en-tetes peuvent etre en francais, arabe, anglais ou abreges.

En-tetes du fichier Excel:
{headers}

Exemples de donnees (3 premieres lignes):
{sample_rows}

Champs canoniques disponibles:
{canonical_fields}

REGLES:
- Si un en-tete correspond a un champ canonique, retourne le mapping.
- Si un en-tete ne correspond a rien, mets null.
- Si "project_code" n'est pas dans les en-tetes, cherche un pattern dans "unit_code" pour le deduire.
- Si "block_code" ou "level_code" ne sont pas explicites, cherche dans les codes lots.
- Tente de comprendre les abbreviations: SH=surface habitable, SU=surface utile, etc.
- "N Lot" ou "Num Lot" = unit_code
- "Etg" ou "Niv" = level_code
- "Bloc" ou "Bat" ou "Immeuble" = block_code
- "Type" ou "Typ" = typology
- "Sup Hab" ou "SH" = area_sh
- "PU" ou "Prix Unit" = price_base
- "PT" ou "Prix Total" ou "Montant" = price_total

Retourne UNIQUEMENT un JSON valide (pas de texte avant/apres) avec cette structure:
{{
  "column_mapping": {{
    "nom_colonne_excel": "champ_canonique_ou_null",
    ...
  }},
  "project_code_default": "CODE_PROJET si deduit du contexte, sinon null",
  "block_code_pattern": "description du pattern si deduit des codes lots, sinon null",
  "notes": "observations sur le format detecte"
}}
"""


async def _call_llm_for_mapping(headers: list[str], sample_rows: list[list[Any]]) -> dict:
    """Call LLM to map Excel columns to canonical EDD fields."""
    from app.config import get_settings

    settings = get_settings()

    canonical_desc = "\n".join(f"  - {k}: {v}" for k, v in CANONICAL_FIELDS.items())
    headers_str = ", ".join(f'"{h}"' for h in headers)
    sample_str = "\n".join(str(row) for row in sample_rows[:3])

    prompt = COLUMN_MAPPING_PROMPT.format(
        headers=headers_str,
        sample_rows=sample_str,
        canonical_fields=canonical_desc,
    )

    try:
        from langchain_openai import ChatOpenAI

        llm = ChatOpenAI(
            model=settings.LLM_MODEL,
            api_key=settings.LLM_API_KEY,
            base_url=settings.LLM_API_BASE,
            temperature=0.0,
            max_tokens=2048,
        )
        response = await llm.ainvoke(prompt)
        content = response.content.strip()

        # Extract JSON from response (handle markdown code blocks)
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            content = content.split("```")[1].split("```")[0].strip()

        return json.loads(content)
    except Exception as e:
        logger.warning("llm_mapping_failed", error=str(e))
        return _fallback_mapping(headers)


def _fallback_mapping(headers: list[str]) -> dict:
    """Rule-based fallback when LLM is unavailable."""
    mapping = {}
    header_rules = {
        "project_code": ["projet", "project", "code projet", "code_projet"],
        "block_code": ["bloc", "bat", "batiment", "immeuble", "building", "block"],
        "level_code": ["etage", "niveau", "etg", "niv", "level", "floor", "rdc"],
        "unit_code": ["lot", "n lot", "num lot", "numero lot", "n°lot", "code lot", "unite", "unit", "code", "ref"],
        "typology": ["type", "typ", "typologie", "typology", "designation"],
        "bedrooms": ["chambre", "ch", "bedroom", "nbr ch", "pieces"],
        "bathrooms": ["sdb", "salle de bain", "bathroom"],
        "area_sh": ["sh", "surface hab", "surface habitable", "sup hab", "s.hab", "s hab"],
        "area_su": ["su", "surface utile", "sup utile", "s.utile", "s utile"],
        "area_sb": ["sb", "balcon", "surface balcon", "s balcon"],
        "area_st": ["st", "terrasse", "surface terrasse", "s terrasse"],
        "area_sj": ["sj", "jardin", "surface jardin", "s jardin"],
        "area_sp": ["sp", "parking", "surface parking", "s parking", "place parking"],
        "area_sc": ["sc", "cave", "surface cave", "s cave", "cellier"],
        "exposure": ["exposition", "orientation", "exposure", "orient"],
        "view_class": ["vue", "view", "type vue"],
        "price_base": ["pu", "prix unit", "prix unitaire", "prix/m2", "prix m2"],
        "price_total": ["pt", "prix total", "montant", "prix", "total", "prix ttc"],
        "usage_type": ["usage", "destination", "affectation"],
    }

    for header in headers:
        h_lower = header.lower().strip()
        matched = False
        for canon_field, keywords in header_rules.items():
            for kw in keywords:
                if kw == h_lower or kw in h_lower:
                    mapping[header] = canon_field
                    matched = True
                    break
            if matched:
                break
        if not matched:
            mapping[header] = None

    return {
        "column_mapping": mapping,
        "project_code_default": None,
        "block_code_pattern": None,
        "notes": "Mapping par regles (LLM non disponible)",
    }


def _parse_level_code(raw: Any) -> str:
    """Normalize level code to Lxx format."""
    if raw is None:
        return "L00"
    s = str(raw).strip().upper()
    if s in ("RDC", "RC", "REZ", "REZ-DE-CHAUSSEE"):
        return "L00"
    if s.startswith("SS") or s.startswith("S-"):
        try:
            num = int("".join(c for c in s if c.isdigit()) or "1")
            return f"S{num:02d}"
        except ValueError:
            return "S01"
    if s.startswith("L") and s[1:].isdigit():
        return f"L{int(s[1:]):02d}"
    try:
        num = int(float(s))
        if num < 0:
            return f"S{abs(num):02d}"
        return f"L{num:02d}"
    except (ValueError, TypeError):
        return s[:10]


def _safe_decimal(val: Any) -> float | None:
    """Convert value to float safely."""
    if val is None or val == "" or val == "-":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        # Try removing spaces, commas
        try:
            cleaned = str(val).replace(" ", "").replace(",", ".").replace("\xa0", "")
            return float(cleaned)
        except (ValueError, TypeError):
            return None


def _safe_int(val: Any) -> int | None:
    """Convert value to int safely."""
    f = _safe_decimal(val)
    return int(f) if f is not None else None


async def import_edd_from_excel(
    file_content: bytes,
    filename: str,
    tenant_id: str,
    user_id: str,
    project_code_override: str | None,
    db: AsyncSession,
) -> dict[str, Any]:
    """Import EDD data from an Excel file using AI-powered column mapping.

    Args:
        file_content: Raw Excel file bytes
        filename: Original filename
        tenant_id: Current user's tenant
        user_id: Current user ID
        project_code_override: If set, forces this project code for all rows
        db: Database session

    Returns:
        Import summary dict with stats and warnings
    """
    # Step 1: Read Excel
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return {"status": "error", "message": "Fichier Excel vide ou invalide"}

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"status": "error", "message": "Le fichier doit contenir au moins un en-tete et une ligne de donnees"}

    # Find header row (first non-empty row)
    header_idx = 0
    for i, row in enumerate(rows):
        if any(cell is not None for cell in row):
            header_idx = i
            break

    raw_headers = [str(cell or "").strip() for cell in rows[header_idx]]
    # Remove empty trailing columns
    while raw_headers and raw_headers[-1] == "":
        raw_headers.pop()

    data_rows = []
    for row in rows[header_idx + 1:]:
        row_vals = list(row[:len(raw_headers)])
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row_vals):
            continue
        data_rows.append(row_vals)

    if not data_rows:
        return {"status": "error", "message": "Aucune ligne de donnees trouvee"}

    # Step 2: AI column mapping
    sample = data_rows[:5]
    ai_result = await _call_llm_for_mapping(raw_headers, sample)
    col_mapping = ai_result.get("column_mapping", {})
    default_project = project_code_override or ai_result.get("project_code_default")

    logger.info(
        "edd_excel_mapping",
        filename=filename,
        headers=raw_headers,
        mapping=col_mapping,
        notes=ai_result.get("notes"),
    )

    # Build column index -> canonical field
    col_map: dict[int, str] = {}
    for idx, header in enumerate(raw_headers):
        canon = col_mapping.get(header)
        if canon and canon in CANONICAL_FIELDS:
            col_map[idx] = canon

    # Verify minimum required fields
    mapped_fields = set(col_map.values())
    has_unit_code = "unit_code" in mapped_fields
    has_area = "area_sh" in mapped_fields or "area_su" in mapped_fields

    if not has_unit_code:
        return {
            "status": "error",
            "message": "Impossible d'identifier la colonne 'code lot/unite'. "
                       "Verifiez que votre fichier contient une colonne avec les codes lots.",
            "detected_headers": raw_headers,
            "ai_mapping": col_mapping,
        }

    # Step 3: Process rows
    warnings: list[dict] = []
    created_units = 0
    updated_units = 0
    skipped_units = 0
    project_cache: dict[str, REProject] = {}
    block_cache: dict[str, REBlock] = {}
    level_cache: dict[str, RELevel] = {}

    for row_idx, row in enumerate(data_rows):
        try:
            # Extract values using mapping
            values: dict[str, Any] = {}
            for col_idx, field in col_map.items():
                if col_idx < len(row):
                    values[field] = row[col_idx]

            unit_code = str(values.get("unit_code", "")).strip()
            if not unit_code:
                skipped_units += 1
                continue

            # Resolve project code
            proj_code = str(values.get("project_code", "") or default_project or "IMPORT").strip()
            if not proj_code:
                proj_code = "IMPORT"

            # Resolve block code
            block_code = str(values.get("block_code", "") or "A").strip().upper()

            # Resolve level code
            level_code = _parse_level_code(values.get("level_code", "L00"))

            # Get or create project
            if proj_code not in project_cache:
                proj_r = await db.execute(
                    select(REProject).where(REProject.code == proj_code)
                )
                project = proj_r.scalar_one_or_none()
                if not project:
                    project = REProject(
                        id=str(uuid.uuid4()),
                        tenant_id=tenant_id,
                        code=proj_code,
                        name=proj_code,
                        status="DRAFT",
                    )
                    db.add(project)
                    await db.flush()
                project_cache[proj_code] = project

            project = project_cache[proj_code]

            # Get or create block
            block_key = f"{project.id}:{block_code}"
            if block_key not in block_cache:
                blk_r = await db.execute(
                    select(REBlock).where(
                        REBlock.project_id == project.id,
                        REBlock.code == block_code,
                    )
                )
                block = blk_r.scalar_one_or_none()
                if not block:
                    block = REBlock(
                        id=str(uuid.uuid4()),
                        project_id=project.id,
                        code=block_code,
                        name=block_code,
                    )
                    db.add(block)
                    await db.flush()
                block_cache[block_key] = block

            block = block_cache[block_key]

            # Get or create level
            level_key = f"{block.id}:{level_code}"
            if level_key not in level_cache:
                lvl_r = await db.execute(
                    select(RELevel).where(
                        RELevel.block_id == block.id,
                        RELevel.code == level_code,
                    )
                )
                level = lvl_r.scalar_one_or_none()
                if not level:
                    level = RELevel(
                        id=str(uuid.uuid4()),
                        block_id=block.id,
                        code=level_code,
                    )
                    db.add(level)
                    await db.flush()
                level_cache[level_key] = level

            level = level_cache[level_key]

            # Create or update unit
            unit_r = await db.execute(
                select(REUnit).where(REUnit.code == unit_code)
            )
            existing_unit = unit_r.scalar_one_or_none()

            unit_data = {
                "typology": str(values.get("typology", "")).strip() or None,
                "bedrooms": _safe_int(values.get("bedrooms")),
                "bathrooms": _safe_int(values.get("bathrooms")),
                "area_sh": _safe_decimal(values.get("area_sh")),
                "area_su": _safe_decimal(values.get("area_su")),
                "area_sb": _safe_decimal(values.get("area_sb")),
                "area_st": _safe_decimal(values.get("area_st")),
                "area_sj": _safe_decimal(values.get("area_sj")),
                "area_sp": _safe_decimal(values.get("area_sp")),
                "area_sc": _safe_decimal(values.get("area_sc")),
                "exposure": str(values.get("exposure", "")).strip().upper() or None,
                "view_class": str(values.get("view_class", "")).strip().lower() or None,
                "price_base": _safe_decimal(values.get("price_base")),
                "price_total": _safe_decimal(values.get("price_total")),
                "usage_type": str(values.get("usage_type", "")).strip() or None,
            }

            if existing_unit:
                if existing_unit.edd_state == EddState.FROZEN.value:
                    warnings.append({
                        "row": row_idx + header_idx + 2,
                        "unit_code": unit_code,
                        "message": "Unite gelee (FROZEN) — ignoree",
                    })
                    skipped_units += 1
                    continue

                # Update non-null values only
                for field, val in unit_data.items():
                    if val is not None:
                        setattr(existing_unit, field, val)
                updated_units += 1
            else:
                unit = REUnit(
                    id=str(uuid.uuid4()),
                    level_id=level.id,
                    code=unit_code,
                    edd_state=EddState.DRAFT.value,
                    **{k: v for k, v in unit_data.items() if v is not None},
                )
                db.add(unit)
                created_units += 1

        except Exception as e:
            warnings.append({
                "row": row_idx + header_idx + 2,
                "message": f"Erreur: {str(e)}",
            })
            skipped_units += 1
            continue

    await db.flush()

    return {
        "status": "success",
        "filename": filename,
        "total_rows": len(data_rows),
        "created": created_units,
        "updated": updated_units,
        "skipped": skipped_units,
        "warnings": warnings,
        "ai_mapping": col_mapping,
        "ai_notes": ai_result.get("notes"),
        "detected_headers": raw_headers,
        "mapped_fields": list(mapped_fields),
    }
